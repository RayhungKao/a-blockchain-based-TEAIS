import json
import os
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from utils.hashing import generate_hashes
from utils.homomorphic_encryption import generate_and_store_keys, get_keys, encrypt_list, decrypt_list, serialize_encrypted_list, deserialize_encrypted_list

# Constants
EXCEL_FILE_NAME = 'at.xlsx'
SHEET_NAME = 'AT_list'
COLUMNS_TO_HASH = ['#', 'seller', 'buyer', 'date', 'quantity', 'unit', 'unit-price', 'amount', 'item-description']
HASHED_COLUMN_NAME = 'hashed-AT-info'
AMOUNT_COLUMN = 'amount'

# Load environment variables from .env file
load_dotenv()

# Retrieve keys from .env file or generate them if they don't exist
public_key, private_key = get_keys()

# Print the public and private keys for development usage
print("Public Key n:", str(public_key.n))
print("Private Key p:", str(private_key.p))
print("Private Key q:", str(private_key.q))

# Read product amounts from Excel file
df = pd.read_excel(EXCEL_FILE_NAME, sheet_name=SHEET_NAME)
print("DataFrame contents:\n", df)
print("DataFrame columns:", df.columns)

# Generate hash for each row in the specified columns
if all(col in df.columns for col in COLUMNS_TO_HASH):
    df[HASHED_COLUMN_NAME] = generate_hashes(df, COLUMNS_TO_HASH)
    print("DataFrame with hashes:\n", df)
else:
    print("One or more specified columns are missing in the DataFrame.")

# Load the existing workbook
workbook = load_workbook(EXCEL_FILE_NAME)
if SHEET_NAME in workbook.sheetnames:
    sheet = workbook[SHEET_NAME]
else:
    print(f"Sheet '{SHEET_NAME}' does not exist in the workbook.")
    exit()

# Find the column index for 'Hashed_AT_info'
hashed_column_index = None
for col in sheet.iter_cols(1, sheet.max_column):
    if col[0].value == HASHED_COLUMN_NAME:
        hashed_column_index = col[0].column_letter
        break

if hashed_column_index is None:
    print(f"Column '{HASHED_COLUMN_NAME}' does not exist in the sheet '{SHEET_NAME}'.")
    exit()

# Write the 'Hashed_AT_info' column to the sheet
for index, value in enumerate(df[HASHED_COLUMN_NAME], start=2):  # Assuming the first row is the header
    sheet[f'{hashed_column_index}{index}'] = value

# Save the workbook
workbook.save(EXCEL_FILE_NAME)
print(f"Updated DataFrame written to {EXCEL_FILE_NAME} in sheet '{SHEET_NAME}'")

# Verify the column name and handle empty DataFrame or missing column
if AMOUNT_COLUMN in df.columns and not df.empty:
    product_amounts = df[AMOUNT_COLUMN].tolist()
else:
    print(f"The column '{AMOUNT_COLUMN}' does not exist or the Excel file is empty. Using dummy array.")
    product_amounts = [10, 20, 30, 40, 50]  # Dummy array

print("product_amounts:", product_amounts)

# Encrypt the list
encrypted_amounts = encrypt_list(public_key, product_amounts)

# Serialize the encrypted amounts
encrypted_amounts_json = serialize_encrypted_list(encrypted_amounts)

# Print the JSON representation of the encrypted amounts
print("Encrypted amounts JSON:", encrypted_amounts_json)

# Deserialize the encrypted amounts
encrypted_amounts_restored = deserialize_encrypted_list(public_key, encrypted_amounts_json)

# Decrypt the list
decrypted_amounts = decrypt_list(private_key, encrypted_amounts_restored)

# Encrypt the sum of the original amounts
encrypted_sum = sum(encrypted_amounts_restored)

# Decrypt the encrypted sum
decrypted_sum = private_key.decrypt(encrypted_sum)

# Print the original, encrypted, and decrypted lists
print("Original amounts:", product_amounts)
print("Encrypted amounts:", encrypted_amounts_restored)
print("Decrypted amounts:", decrypted_amounts)

# Print the sum of the original amounts and the decrypted sum
print("Sum of original amounts:", sum(product_amounts))
print("Decrypted sum of encrypted amounts:", decrypted_sum)